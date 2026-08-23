"""
调试全员互评问题 - 所有人打分都一样，但输出结果不同
"""
import openpyxl
import os
import sys

# 检查全员互评文件夹中的评分数据
test_dir = '全员互评'
files = sorted([f for f in os.listdir(test_dir) if f.endswith('.xlsx')])

print("=" * 80)
print("全员互评文件夹 - 评分数据检查")
print("=" * 80)
print()

# 检查每个评分人的评分
for file in files:
    filepath = os.path.join(test_dir, file)
    wb = openpyxl.load_workbook(filepath)
    
    # 找到员工评价表
    sheet_name = None
    for name in wb.sheetnames:
        if '员工评价表' in name or '评价表' in name:
            sheet_name = name
            break
    
    if not sheet_name:
        print(f"⚠️  {file} - 找不到评价表")
        continue
    
    ws = wb[sheet_name]
    
    # 获取评分人信息
    evaluator_name = ws.cell(2, 6).value  # 第2行第6列
    evaluator_group = ws.cell(2, 7).value  # 第2行第7列
    
    print(f"📄 {file}")
    print(f"   评分人: {evaluator_name} ({evaluator_group})")
    print(f"   评分数据:")
    
    # 遍历被评分人
    count = 0
    for row_idx, row in enumerate(ws.iter_rows(min_row=4, values_only=True), start=4):
        if not row[1]:  # 考评对象为空，停止
            break
        
        evaluated_name = row[1]
        evaluated_group = row[2]
        score1 = row[4]  # 第一维度评分
        score2 = row[5]  # 第二维度评分
        score3 = row[6]  # 第三维度评分
        
        count += 1
        if count <= 5:  # 只显示前5个
            print(f"      {evaluated_name:8} ({evaluated_group:8}): {score1}, {score2}, {score3}")
    
    print(f"   总共: {count} 个被评分人")
    print()

print("=" * 80)
print("预期：所有评分人都给所有被评分人打相同的分数")
print("问题：最终输出表中的分数不同，有些为零")
print("=" * 80)

