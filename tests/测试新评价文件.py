#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import sys
import shutil
import openpyxl
import main

print('=' * 100)
print('测试新评价文件 - 测试文档_20251211')
print('=' * 100)
print()

# 步骤1：清理旧的输出文件
print('✓ 步骤1：清理旧的输出文件')
print('-' * 100)
output_file = '2025年度架构与技术管理处人员评价汇总表.xlsx'
if os.path.exists(output_file):
    os.remove(output_file)
    print(f'  已删除旧文件：{output_file}')
print()

# 步骤2：准备测试数据
print('✓ 步骤2：准备测试数据')
print('-' * 100)
config_file = '人员配置表2025.xlsx'
test_dir = '测试文档_20251211'

if not os.path.exists(config_file):
    print(f'  ✗ 配置文件不存在：{config_file}')
    sys.exit(1)

if not os.path.exists(test_dir):
    print(f'  ✗ 测试目录不存在：{test_dir}')
    sys.exit(1)

merits_files = [os.path.join(test_dir, f) for f in os.listdir(test_dir) if f.endswith('.xlsx')]
print(f'  ✓ 配置文件：{config_file}')
print(f'  ✓ 评价文件数量：{len(merits_files)}')
print()

# 步骤3：执行计算
print('✓ 步骤3：执行计算')
print('-' * 100)

# 初始化
main.init(auto_scan=False)

# 设置全局变量
main.CONFIG_FILE_PATH = os.path.abspath(config_file)
main.MERITS_FILE_PATH_LIST.clear()
main.MERITS_FILE_PATH_LIST.extend([os.path.abspath(f) for f in merits_files])
main.MERITS_MAP.clear()

# 执行计算
main.init_config()
main.check_file()
main.computed()
main.out()

print('  ✓ 计算完成')
print()

# 步骤4：验证输出文件
print('✓ 步骤4：验证输出文件')
print('-' * 100)

if not os.path.exists(output_file):
    print(f'  ✗ 输出文件不存在')
    sys.exit(1)

wb = openpyxl.load_workbook(output_file)
ws = wb['评分汇总']

record_count = ws.max_row - 1  # 减去表头
print(f'  ✓ 输出文件已生成')
print(f'  ✓ 员工记录数：{record_count}')
print()

# 步骤5：显示前15条记录
print('✓ 步骤5：前15条记录预览')
print('-' * 100)
print(f'{"排名":<6} {"姓名":<12} {"层级":<10} {"总分":<8}')
print('-' * 100)

for i, row in enumerate(ws.iter_rows(min_row=2, max_row=16, values_only=True), 1):
    rank, group, name, level, director, vice_director, group_leader, other, total = row
    print(f'{rank:<6} {name:<12} {level:<10} {total:<8.2f}')

print()
print('=' * 100)
print('✓ 测试完成！')
print('=' * 100)
print()
print(f'输出文件：{output_file}')
print(f'员工总数：{record_count}')

