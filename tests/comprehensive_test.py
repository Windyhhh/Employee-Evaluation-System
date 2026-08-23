"""
全面验证测试：模拟所有人给所有人打满分
生成完整的测试数据，确保每个人都有完整的评分来源
"""
import sys
import os
import openpyxl
from openpyxl import Workbook
sys.path.insert(0, '.')

# 清理之前的模块缓存
for mod in list(sys.modules.keys()):
    if 'main' in mod:
        del sys.modules[mod]

# 先读取人员配置
print("=== 读取人员配置 ===")
config_wb = openpyxl.load_workbook('人员配置表2025.xlsx')
config_ws = config_wb['人员配置']

# 收集所有人员信息
all_members = []
for row in config_ws.iter_rows(min_row=2, values_only=True):
    if row[2]:  # 考评对象不为空
        member = {
            'group': row[0],
            'group_leader': row[1],
            'name': row[2],
            'level': row[3]
        }
        # 去重（王岳珑和王丽丽有两行）
        if not any(m['name'] == member['name'] for m in all_members):
            all_members.append(member)
            print(f"  {member['name']} - {member['level']} ({member['group']})")

print(f"\n共 {len(all_members)} 人")

# 创建测试评价文件（每个人都给所有人打满分）
print("\n=== 生成测试评价文件 ===")
test_files = []

for evaluator in all_members:
    filename = f"测试评价_{evaluator['name']}.xlsx"
    test_files.append(filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = '述职评分表'
    
    # 表头
    ws['A1'] = '序号'
    ws['B1'] = '考评对象'
    ws['C1'] = '职能组'
    ws['D1'] = '员工层级'
    ws['E1'] = '职级'
    ws['F1'] = '敏捷自驱'
    ws['G1'] = '追求卓越'
    ws['H1'] = '超越自我'
    
    # 评分人信息
    ws['E2'] = '评分人：'
    ws['F2'] = evaluator['name']
    
    # 填写评分（从第4行开始）
    row_num = 4
    for evaluated in all_members:
        if evaluated['name'] != evaluator['name']:  # 不给自己打分
            ws[f'A{row_num}'] = row_num - 3
            ws[f'B{row_num}'] = evaluated['name']
            ws[f'C{row_num}'] = evaluated['group']
            ws[f'D{row_num}'] = evaluated['level']
            ws[f'E{row_num}'] = 'T9'
            ws[f'F{row_num}'] = '非常满足'
            ws[f'G{row_num}'] = '非常满足'
            ws[f'H{row_num}'] = '非常满足'
            row_num += 1
    
    wb.save(filename)

print(f"生成了 {len(test_files)} 个测试文件")

# 运行评分计算
print("\n=== 运行评分计算 ===")

# 删除旧输出文件
output_file = '2025年度架构与技术管理处人员评价汇总表.xlsx'
if os.path.exists(output_file):
    os.remove(output_file)

from main import init, init_config, computed, out, MERITS_MAP, MERITS_FILE_PATH_LIST

# 清空并重新添加测试文件
MERITS_FILE_PATH_LIST.clear()
for f in test_files:
    MERITS_FILE_PATH_LIST.append(os.path.abspath(f))

init(auto_scan=False)
init_config()
computed()
out()

# 检查结果
print("\n=== 验证结果 ===")
result_wb = openpyxl.load_workbook(output_file)
result_ws = result_wb['评分汇总']

print(f"{'排名':<4} {'姓名':<8} {'层级':<8} {'处长':<8} {'副处长':<8} {'组长':<8} {'互评':<8} {'总分':<8} {'状态'}")
print("-" * 90)

errors = []
for row in result_ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break
    rank, group, name, level, director, vice_dir, group_leader, other, total = row
    
    # 检查各项分数
    status_parts = []
    
    if director is None or director == 0:
        status_parts.append("处长=0")
    if vice_dir is None or vice_dir == 0:
        status_parts.append("副处长=0")
    if level == '普通员工' and (group_leader is None or group_leader == 0 or group_leader == ''):
        status_parts.append("组长=0")
    if other is None or other == 0:
        status_parts.append("互评=0")
    
    if total is not None and abs(total - 100) < 0.01:
        status = "✓"
    else:
        status = f"✗ {total:.2f}" if total else "✗ 0"
        if status_parts:
            status += f" ({', '.join(status_parts)})"
        errors.append((name, level, total, status_parts))
    
    d_str = f"{director:.2f}" if director else "0.00"
    v_str = f"{vice_dir:.2f}" if vice_dir else "0.00"
    g_str = f"{group_leader:.2f}" if group_leader and group_leader != '' else "-"
    o_str = f"{other:.2f}" if other else "0.00"
    t_str = f"{total:.2f}" if total else "0.00"
    
    print(f"{rank:<4} {name:<8} {level:<8} {d_str:<8} {v_str:<8} {g_str:<8} {o_str:<8} {t_str:<8} {status}")

print()
if errors:
    print(f"❌ 发现 {len(errors)} 个错误:")
    for name, level, total, parts in errors:
        print(f"  - {name}({level}): 总分={total}, 缺失: {parts}")
else:
    print("✅ 所有人得分均为100分，修复成功！")

# 清理测试文件
print("\n=== 清理测试文件 ===")
for f in test_files:
    if os.path.exists(f):
        os.remove(f)
print("清理完成")

