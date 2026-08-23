#!/usr/bin/env python
# -*- coding: utf-8 -*-

import openpyxl

print('=' * 100)
print('检查是否符合任务要求')
print('=' * 100)
print()

# 1. 检查人员配置表
print('✓ 1. 人员配置表检查')
print('-' * 100)

config_file = '人员配置表2025.xlsx'
wb = openpyxl.load_workbook(config_file)
ws = wb['人员配置']

# 统计各角色
roles = {}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[2]:  # 姓名列
        name = row[2]
        level = row[3]
        if level not in roles:
            roles[level] = []
        roles[level].append(name)

print(f'  处长：{roles.get("处长", [])}')
print(f'  副处长：{sorted(roles.get("副处长", []))}')
print(f'  组长：{sorted(roles.get("组长", []))}')
print(f'  小队长：{sorted(roles.get("小队长", []))}')
print(f'  普通员工：{len(roles.get("普通员工", []))}人')

print()

# 2. 检查王岳珑和王丽丽的配置
print('✓ 2. 王岳珑和王丽丽的配置检查')
print('-' * 100)

print(f'  王岳珑是副处长：{"王岳珑" in roles.get("副处长", [])}')
print(f'  王岳珑是组长：{"王岳珑" in roles.get("组长", [])}')
print(f'  王丽丽是副处长：{"王丽丽" in roles.get("副处长", [])}')
print(f'  王丽丽是组长：{"王丽丽" in roles.get("组长", [])}')

print()

# 3. 检查输出文件
print('✓ 3. 输出文件检查')
print('-' * 100)

output_file = '2025年度架构与技术管理处人员评价汇总表.xlsx'
wb_out = openpyxl.load_workbook(output_file)
ws_out = wb_out['评分汇总']

record_count = ws_out.max_row - 1
print(f'  输出记录数：{record_count}')

# 检查是否包含王岳珑和王丽丽
found_yuanlong = False
found_lili = False
for row in ws_out.iter_rows(min_row=2, max_row=ws_out.max_row, values_only=True):
    if row[2] == '王岳珑':
        found_yuanlong = True
    if row[2] == '王丽丽':
        found_lili = True

print(f'  王岳珑在输出中：{found_yuanlong}（应该为False）')
print(f'  王丽丽在输出中：{found_lili}（应该为False）')

print()

# 4. 检查权重规则
print('✓ 4. 权重规则检查')
print('-' * 100)

# 检查一个组长的评分
for row in ws_out.iter_rows(min_row=2, max_row=10, values_only=True):
    if row[3] == '组长':
        print(f'  组长示例：{row[2]}')
        print(f'    处长评价：{row[4]}')
        print(f'    副处长评价：{row[5]}')
        print(f'    组长评价：{row[6]}（应该为空）')
        print(f'    其他人员互评：{row[7]}')
        print(f'    总分：{row[8]}')
        break

print()

# 检查一个普通员工的评分
for row in ws_out.iter_rows(min_row=2, max_row=20, values_only=True):
    if row[3] == '普通员工':
        print(f'  普通员工示例：{row[2]}')
        print(f'    处长评价：{row[4]}')
        print(f'    副处长评价：{row[5]}')
        print(f'    组长评价：{row[6]}')
        print(f'    其他人员互评：{row[7]}')
        print(f'    总分：{row[8]}')
        break

print()
print('=' * 100)

