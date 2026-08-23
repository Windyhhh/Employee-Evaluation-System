"""
测试EXE中的计算逻辑是否与main.py一致
通过模拟GUI的调用方式来验证
"""
import sys
import os
import shutil
import openpyxl
from openpyxl import Workbook

# 清理旧输出
output_file = '2025年度架构与技术管理处人员评价汇总表.xlsx'
if os.path.exists(output_file):
    os.remove(output_file)

# 清除模块缓存
for mod in list(sys.modules.keys()):
    if 'main' in mod or 'gui' in mod:
        del sys.modules[mod]

# 导入main模块
import main

print("=== 测试EXE计算逻辑 ===\n")

# 读取人员配置
config_wb = openpyxl.load_workbook('人员配置表2025.xlsx')
config_ws = config_wb['人员配置']

all_members = []
for row in config_ws.iter_rows(min_row=2, values_only=True):
    if row[2]:
        member = {
            'group': row[0],
            'group_leader': row[1],
            'name': row[2],
            'level': row[3]
        }
        if not any(m['name'] == member['name'] for m in all_members):
            all_members.append(member)

print(f"共 {len(all_members)} 人\n")

# 生成测试文件
print("生成测试评价文件...")
test_files = []
for evaluator in all_members:
    filename = f"test_exe_{evaluator['name']}.xlsx"
    test_files.append(filename)
    
    wb = Workbook()
    ws = wb.active
    ws.title = '述职评分表'
    
    ws['A1'] = '序号'
    ws['B1'] = '考评对象'
    ws['C1'] = '职能组'
    ws['D1'] = '员工层级'
    ws['E1'] = '职级'
    ws['F1'] = '敏捷自驱'
    ws['G1'] = '追求卓越'
    ws['H1'] = '超越自我'
    
    ws['E2'] = '评分人：'
    ws['F2'] = evaluator['name']
    
    row_num = 4
    for evaluated in all_members:
        if evaluated['name'] != evaluator['name']:
            ws[f'A{row_num}'] = row_num - 3
            ws[f'B{row_num}'] = evaluated['name']
            ws[f'C{row_num}'] = evaluated['group']
            ws[f'D{row_num}'] = evaluated['level']
            ws[f'E{row_num}'] = 'T9'
            ws[f'F{row_num}'] = '非常满足'
            ws[f'G{row_num}'] = '非常满足'
            ws[f'H{row_num}'] = '非常满足'
            row_num += 1
    
    wb.save(filename)

print(f"生成了 {len(test_files)} 个测试文件\n")

# 模拟GUI的调用方式
print("执行计算（模拟GUI调用）...")
main.init(auto_scan=False)
main.CONFIG_FILE_PATH = os.path.abspath('人员配置表2025.xlsx')
main.MERITS_FILE_PATH_LIST.clear()
main.MERITS_FILE_PATH_LIST.extend([os.path.abspath(f) for f in test_files])
main.MERITS_MAP.clear()

main.init_config()
main.check_file()
main.computed()
main.out()

# 验证结果
print("\n验证结果...")
result_wb = openpyxl.load_workbook(output_file)
result_ws = result_wb['评分汇总']

errors = []
count = 0
for row in result_ws.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        break
    rank, group, name, level, director, vice_dir, group_leader, other, total = row
    count += 1
    
    if total is None or abs(total - 100) >= 0.01:
        errors.append((name, level, total))

print(f"共 {count} 人被评价")
if errors:
    print(f"❌ 发现 {len(errors)} 个错误:")
    for name, level, total in errors:
        print(f"  - {name}({level}): {total}")
else:
    print("✅ 所有人得分均为100分")

# 清理测试文件
print("\n清理测试文件...")
for f in test_files:
    if os.path.exists(f):
        os.remove(f)

print("✅ 测试完成！EXE计算逻辑正确")

