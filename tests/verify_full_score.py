"""
验证全员满分时最终得分应为100
"""
import sys
import os
import shutil
import openpyxl
sys.path.insert(0, '.')

# 清理旧的输出文件
output_file = '2025年度架构与技术管理处人员评价汇总表.xlsx'
if os.path.exists(output_file):
    os.remove(output_file)
    print(f"已删除旧文件: {output_file}")

# 复制测试文件到主目录并修改为全员满分
test_dir = '测试文档_20251211'
files = sorted([f for f in os.listdir(test_dir) if f.endswith('.xlsx')])

print(f"\n处理 {len(files)} 个测试文件（全部改为非常满足）...")
for file in files:
    src = os.path.join(test_dir, file)
    dst = file
    
    # 加载并修改评分为全员满分
    wb = openpyxl.load_workbook(src)
    ws = wb.active
    
    # 修改所有评分为"非常满足"
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        if row[1].value is None:  # 考评人为空，停止
            break
        # 评分在第6、7、8列（索引5、6、7）
        row[5].value = '非常满足'
        row[6].value = '非常满足'
        row[7].value = '非常满足'
    
    wb.save(dst)
    print(f"  OK: {file}")

# 重新导入main模块（确保使用最新代码）
if 'main' in sys.modules:
    del sys.modules['main']

from main import init, init_config, computed, out, MERITS_MAP, ROLE_RATIO, LEADER_ROLE_RATIO

print("\n初始化配置...")
init()
init_config()

print(f"普通员工权重: {ROLE_RATIO}")
print(f"组长/小队长权重: {LEADER_ROLE_RATIO}")

print("\n计算评分...")
computed()

print("\n输出结果...")
out()

print(f"\n✅ 完成！输出文件: {output_file}")

# 检查输出结果
wb = openpyxl.load_workbook(output_file)
ws = wb['评分汇总']

print("\n=== 输出结果 ===")
print(f"{'排名':<4} {'姓名':<8} {'层级':<8} {'处长':<8} {'副处长':<8} {'组长':<8} {'互评':<8} {'总分':<8}")
print("-" * 80)

has_error = False
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
    if row[0] is None:
        break
    rank, group, name, level, director, vice_dir, group_leader, other, total = row
    
    # 格式化输出
    director_str = f"{director:.2f}" if director else "0.00"
    vice_dir_str = f"{vice_dir:.2f}" if vice_dir else "0.00"
    group_leader_str = f"{group_leader:.2f}" if group_leader else "-"
    other_str = f"{other:.2f}" if other else "0.00"
    total_str = f"{total:.2f}" if total else "0.00"
    
    # 检查是否为100分
    status = "✓" if abs(total - 100) < 0.01 else "✗ ERROR"
    if abs(total - 100) >= 0.01:
        has_error = True
    
    print(f"{rank:<4} {name:<8} {level:<8} {director_str:<8} {vice_dir_str:<8} {group_leader_str:<8} {other_str:<8} {total_str:<8} {status}")

print()
if has_error:
    print("❌ 存在得分不为100的情况！")
else:
    print("✅ 所有人得分均为100分！")

# 清理测试文件
print("\n清理测试文件...")
for file in files:
    if os.path.exists(file):
        os.remove(file)
print("清理完成")

