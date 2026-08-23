'''
Author: yl_li
Date: 2023-12-11
LastEditors: yl_li
LastEditTime: 2024-12-21
description: 员工评价管理系统 - 2025年版本
'''
import os
import logging
import sys
import openpyxl
import copy

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s]: %(message)s',
    handlers=[logging.StreamHandler()]
)

WORKSPEACE_DIR = ''
CONFIG_FILE_PATH = ''
RESULT_FILE_PATH = ''
MERITS_FILE_PATH_LIST = []
MERITS_MAP = {}  # {考评人名字: Member对象}
CONFIG_MEMBER_SHEET_NAME = '人员配置'
CONFIG_MEMBER_RATIO_SHEET_NAME = '分数配置'
CONFIG_ROLE_RATIO_SHEET_NAME = '角色配置'
MEMBER_SHEET_NAME = '述职评分表'

# 评分标准映射
MERITS_RATIO_1 = {'非常满足': 0, '比较满足': 0, '基本满足': 0, '不满足': 0}
MERITS_RATIO_2 = {'非常满足': 0, '比较满足': 0, '基本满足': 0, '不满足': 0}
MERITS_RATIO_3 = {'非常满足': 0, '比较满足': 0, '基本满足': 0, '不满足': 0}

# 角色权重：处长评价、本组组长评价、其他组组长评价、本组人员互评、其他组人员互评
ROLE_RATIO = (0.2, 0.4, 0.2, 0.1, 0.1)
# 组长权重：处长评价、其他组组长评价、本组人员互评、他组人员互评
LEADER_ROLE_RATIO = (0.4, 0.3, 0.2, 0.1)

# 处长和职能组组长列表
DIRECTOR = []  # 处长
GROUP_LEADERS = {}  # {职能组: 组长名字}
STAFF_LEADERS = {}  # {人员名字: 是否为组长}


class Score:
    """评分类 - 用于存储和计算平均分"""
    def __init__(self) -> None:
        self.score = 0
        self.num = 0

    def add(self, s) -> None:
        self.score += s
        self.num += 1

    def get_score(self):
        if self.num == 0:
            return 0
        else:
            return self.score / self.num

    def __str__(self):
        return f"score: {self.score}, num: {self.num}"


class Member:
    """人员类 - 存储人员信息和评分"""
    def __init__(self, config) -> None:
        # config: [职能组, 职能组组长, 考评人, 管理层级标注]
        self.group = config[0]  # 职能组
        self.group_leader = config[1]  # 职能组组长
        self.name = config[2]  # 考评人
        self.level = config[3]  # 管理层级标注
        self.has_merits_excel = False

        # 5个评分来源：处长、本组组长、其他组组长、本组人员、其他组人员
        # 3个评分维度：敏捷自驱、追求卓越、超越自我
        # 注意：对于组长，索引含义不同
        # 普通员工: [0]处长 [1]本组组长 [2]其他组组长 [3]本组人员 [4]其他组人员
        # 组长: [0]处长 [1]其他组组长 [2]本组人员 [3]他组人员 [4]未使用
        self.score1 = [Score(), Score(), Score(), Score(), Score()]  # 敏捷自驱
        self.score2 = [Score(), Score(), Score(), Score(), Score()]  # 追求卓越
        self.score3 = [Score(), Score(), Score(), Score(), Score()]  # 超越自我

    def __str__(self):
        return f"name: {self.name}, group: {self.group}, leader: {self.group_leader}"

    def has_merits(self):
        self.has_merits_excel = True

    def return_excel(self):
        """返回最终的评分结果"""
        # 计算每个评分来源的平均分
        director_score = (self.score1[0].get_score() + self.score2[0].get_score() + self.score3[0].get_score()) / 3

        # 检查是否为组长，使用不同的索引含义
        if self.name in STAFF_LEADERS:
            # 组长的索引含义：
            # [0] 处长评价
            # [1] 其他组组长评价
            # [2] 本组人员评价（包括组长自评）
            # [3] 他组人员评价
            other_leader_score = (self.score1[1].get_score() + self.score2[1].get_score() + self.score3[1].get_score()) / 3
            group_member_score = (self.score1[2].get_score() + self.score2[2].get_score() + self.score3[2].get_score()) / 3
            other_member_score = (self.score1[3].get_score() + self.score2[3].get_score() + self.score3[3].get_score()) / 3

            # 组长权重：处长评价(0.4)、其他组组长评价(0.3)、本组人员评价(0.2)、他组人员评价(0.1)
            total_score = (director_score * LEADER_ROLE_RATIO[0] +
                          other_leader_score * LEADER_ROLE_RATIO[1] +
                          group_member_score * LEADER_ROLE_RATIO[2] +
                          other_member_score * LEADER_ROLE_RATIO[3])

            # 返回格式：[group, name, level, None, 处长, 本组组长(0), 其他组组长, 本组人员, 他组人员, 总分]
            return [self.group, self.name, self.level, None,
                    director_score, 0, other_leader_score,
                    group_member_score, other_member_score, total_score]
        else:
            # 普通员工的索引含义：
            # [0] 处长评价
            # [1] 本组组长评价
            # [2] 其他组组长评价
            # [3] 本组人员互评
            # [4] 其他组人员互评
            group_leader_score = (self.score1[1].get_score() + self.score2[1].get_score() + self.score3[1].get_score()) / 3
            other_leader_score = (self.score1[2].get_score() + self.score2[2].get_score() + self.score3[2].get_score()) / 3
            group_member_score = (self.score1[3].get_score() + self.score2[3].get_score() + self.score3[3].get_score()) / 3
            other_member_score = (self.score1[4].get_score() + self.score2[4].get_score() + self.score3[4].get_score()) / 3

            # 普通员工权重：处长评价、本组组长评价、其他组组长评价、本组人员互评、其他组人员互评
            total_score = (director_score * ROLE_RATIO[0] +
                          group_leader_score * ROLE_RATIO[1] +
                          other_leader_score * ROLE_RATIO[2] +
                          group_member_score * ROLE_RATIO[3] +
                          other_member_score * ROLE_RATIO[4])

            return [self.group, self.name, self.level, None,
                    director_score, group_leader_score, other_leader_score,
                    group_member_score, other_member_score, total_score]


def init(auto_scan=True):
    """初始化 - 获取工作目录和配置文件路径"""
    if getattr(sys, 'frozen', False):
        exe_path = sys.argv[0]
        exe_absolute_path = os.path.abspath(exe_path)
        current_directory = os.path.dirname(exe_absolute_path)
    else:
        current_directory = os.path.dirname(os.path.abspath(__file__))

    global WORKSPEACE_DIR
    WORKSPEACE_DIR = current_directory
    logging.info(f"当前文件夹位置: {WORKSPEACE_DIR}")

    if auto_scan:
        files_and_folders = os.listdir(current_directory)
        for file in files_and_folders:
            file_path = os.path.join(current_directory, file)
            if os.path.isfile(file_path) and (os.path.splitext(file_path)[1].lower() in ['.xls', '.xlsx']):
                if '配置表' not in file and '汇总表' not in file:
                    MERITS_FILE_PATH_LIST.append(file_path)

    global CONFIG_FILE_PATH
    if not CONFIG_FILE_PATH:
        config_file_names = ['人员配置表.xlsx', '人员配置表2025.xlsx']
        for config_name in config_file_names:
            potential_path = os.path.join(current_directory, config_name)
            if os.path.isfile(potential_path):
                CONFIG_FILE_PATH = potential_path
                break

        if not CONFIG_FILE_PATH:
            CONFIG_FILE_PATH = os.path.join(current_directory, '人员配置表.xlsx')

    logging.info(f"配置文件位置: {CONFIG_FILE_PATH}")
    
    global RESULT_FILE_PATH
    if not RESULT_FILE_PATH:
        RESULT_FILE_PATH = os.path.join(
            current_directory, '2025年度架构与技术管理处人员评价汇总表.xlsx')
    logging.info(f"输出文件位置: {RESULT_FILE_PATH}")


def init_config():
    """读取配置文件"""
    if os.path.isfile(CONFIG_FILE_PATH):
        config_handler = openpyxl.load_workbook(CONFIG_FILE_PATH)

        # 读取人员配置
        config_sheet_handler = config_handler[CONFIG_MEMBER_SHEET_NAME]
        global DIRECTOR, GROUP_LEADERS
        DIRECTOR.clear()
        GROUP_LEADERS.clear()

        for row in config_sheet_handler.iter_rows(min_row=2, max_row=config_sheet_handler.max_row, values_only=True):
            if row[2]:  # 考评人不为空
                MERITS_MAP[row[2]] = Member(row)

                # 建立职能组和组长的映射
                if row[0] and row[1]:  # 职能组和职能组组长都不为空
                    GROUP_LEADERS[row[0]] = row[1]

        # 读取分数配置
        config_sheet_handler = config_handler[CONFIG_MEMBER_RATIO_SHEET_NAME]
        global MERITS_RATIO_1, MERITS_RATIO_2, MERITS_RATIO_3
        MERITS_RATIO_1 = {'非常满足': config_sheet_handler["B2"].value, '比较满足': config_sheet_handler["C2"].value,
                          '基本满足': config_sheet_handler["D2"].value, '不满足': config_sheet_handler["E2"].value}
        MERITS_RATIO_2 = {'非常满足': config_sheet_handler["B3"].value, '比较满足': config_sheet_handler["C3"].value,
                          '基本满足': config_sheet_handler["D3"].value, '不满足': config_sheet_handler["E3"].value}
        MERITS_RATIO_3 = {'非常满足': config_sheet_handler["B4"].value, '比较满足': config_sheet_handler["C4"].value,
                          '基本满足': config_sheet_handler["D4"].value, '不满足': config_sheet_handler["E4"].value}

        # 读取角色配置
        config_sheet_handler = config_handler[CONFIG_ROLE_RATIO_SHEET_NAME]
        global ROLE_RATIO
        ROLE_RATIO = (config_sheet_handler["A2"].value, config_sheet_handler["B2"].value,
                      config_sheet_handler["C2"].value, config_sheet_handler["D2"].value,
                      config_sheet_handler["E2"].value)

        logging.info(f"角色权重: 处长={ROLE_RATIO[0]}, 本组组长={ROLE_RATIO[1]}, 其他组组长={ROLE_RATIO[2]}, 本组人员={ROLE_RATIO[3]}, 其他组人员={ROLE_RATIO[4]}")

        # 识别处长和组长（管理层级标注为"处长"或"组长"的人）
        global STAFF_LEADERS
        STAFF_LEADERS.clear()
        for name, member in MERITS_MAP.items():
            if member.level == '处长':
                DIRECTOR.append(name)
            elif member.level == '组长':
                STAFF_LEADERS[name] = True

        logging.info(f"处长: {DIRECTOR}")
        logging.info(f"职能组组长: {GROUP_LEADERS}")
        logging.info(f"员工组长: {list(STAFF_LEADERS.keys())}")


def check_file():
    """检查评价文件"""
    for merits_file in MERITS_FILE_PATH_LIST:
        file_name = os.path.splitext(os.path.basename(merits_file))[0]
        excel_handler = openpyxl.load_workbook(merits_file)

        # 遍历所有工作表
        for sheet_name in excel_handler.sheetnames:
            # 跳过非评分表的工作表
            if sheet_name not in [MEMBER_SHEET_NAME, '处长评分', '述职评分表', '员工评价表']:
                continue

            try:
                sheet_handler = excel_handler[sheet_name]
            except:
                continue

            # 获取评分人信息（从第2行第5列）
            evaluator_name = sheet_handler.cell(2, 5).value
            if evaluator_name:
                if evaluator_name in MERITS_MAP:
                    MERITS_MAP[evaluator_name].has_merits()
                else:
                    logging.warning(f'【{evaluator_name}】不在人员配置表中')


def computed():
    """计算评分"""
    for merits_file in MERITS_FILE_PATH_LIST:
        logging.debug(f"当前文件: {merits_file}")
        excel_handler = openpyxl.load_workbook(merits_file)

        # 遍历所有工作表
        for sheet_name in excel_handler.sheetnames:
            # 跳过非评分表的工作表
            if sheet_name not in [MEMBER_SHEET_NAME, '处长评分', '述职评分表', '员工评价表']:
                continue

            try:
                sheet_handler = excel_handler[sheet_name]
            except:
                continue

            # 获取评分人信息（从第2行第5列）
            evaluator_name = sheet_handler.cell(2, 5).value
            if not evaluator_name:
                logging.warning(f'工作表【{sheet_name}】中没有填写评分人')
                continue

            if evaluator_name not in MERITS_MAP:
                logging.warning(f'【{evaluator_name}】不在人员配置表中')
                continue

            evaluator = MERITS_MAP[evaluator_name]
            evaluator_group = evaluator.group

            # 遍历被评分人（从第4行开始）
            for row in sheet_handler.iter_rows(min_row=4, values_only=True):
                if not row[1]:  # 考评人为空，停止
                    break

                evaluated_name = row[1]
                evaluated_group = row[2]

                if evaluated_name not in MERITS_MAP:
                    logging.warning(f"【{evaluated_name}】不在人员配置表中，跳过")
                    continue

                # 检查评分数据（第5、6、7列）
                if row[4] is None or row[5] is None or row[6] is None:
                    logging.warning(f"【{evaluator_name}】对【{evaluated_name}】的评分不完整")
                    continue

                evaluated = MERITS_MAP[evaluated_name]

                # 检查被评分人是否为组长，使用不同的权重配置
                is_evaluated_leader = evaluated_name in STAFF_LEADERS

                # 确定评分人的角色
                if evaluator_name in DIRECTOR:
                    # 处长评价
                    evaluated.score1[0].add(MERITS_RATIO_1.get(row[4], 0))
                    evaluated.score2[0].add(MERITS_RATIO_2.get(row[5], 0))
                    evaluated.score3[0].add(MERITS_RATIO_3.get(row[6], 0))
                elif not is_evaluated_leader and evaluator_name == GROUP_LEADERS.get(evaluated_group):
                    # 本组组长评价（评分人是被评分人所在组的组长，且被评分人不是组长）
                    evaluated.score1[1].add(MERITS_RATIO_1.get(row[4], 0))
                    evaluated.score2[1].add(MERITS_RATIO_2.get(row[5], 0))
                    evaluated.score3[1].add(MERITS_RATIO_3.get(row[6], 0))
                elif evaluator_name in GROUP_LEADERS.values() and evaluator_group != evaluated_group:
                    # 其他组组长评价（评分人是其他组的组长）
                    if is_evaluated_leader:
                        # 如果被评分人是组长，使用其他组组长评价（组长的索引1）
                        evaluated.score1[1].add(MERITS_RATIO_1.get(row[4], 0))
                        evaluated.score2[1].add(MERITS_RATIO_2.get(row[5], 0))
                        evaluated.score3[1].add(MERITS_RATIO_3.get(row[6], 0))
                    else:
                        # 普通员工使用其他组组长评价（索引2）
                        evaluated.score1[2].add(MERITS_RATIO_1.get(row[4], 0))
                        evaluated.score2[2].add(MERITS_RATIO_2.get(row[5], 0))
                        evaluated.score3[2].add(MERITS_RATIO_3.get(row[6], 0))
                elif evaluator_group == evaluated_group:
                    # 本组人员互评
                    if is_evaluated_leader:
                        # 如果被评分人是组长，使用本组人员评价（组长的索引2）
                        evaluated.score1[2].add(MERITS_RATIO_1.get(row[4], 0))
                        evaluated.score2[2].add(MERITS_RATIO_2.get(row[5], 0))
                        evaluated.score3[2].add(MERITS_RATIO_3.get(row[6], 0))
                    else:
                        # 普通员工使用本组人员互评（索引3）
                        evaluated.score1[3].add(MERITS_RATIO_1.get(row[4], 0))
                        evaluated.score2[3].add(MERITS_RATIO_2.get(row[5], 0))
                        evaluated.score3[3].add(MERITS_RATIO_3.get(row[6], 0))
                else:
                    # 其他组人员互评
                    if is_evaluated_leader:
                        # 如果被评分人是组长，使用他组人员评价（组长的索引3）
                        evaluated.score1[3].add(MERITS_RATIO_1.get(row[4], 0))
                        evaluated.score2[3].add(MERITS_RATIO_2.get(row[5], 0))
                        evaluated.score3[3].add(MERITS_RATIO_3.get(row[6], 0))
                    else:
                        # 普通员工使用其他组人员互评（索引4）
                        evaluated.score1[4].add(MERITS_RATIO_1.get(row[4], 0))
                        evaluated.score2[4].add(MERITS_RATIO_2.get(row[5], 0))
                        evaluated.score3[4].add(MERITS_RATIO_3.get(row[6], 0))


def out():
    """输出结果"""
    result_file_handler = openpyxl.Workbook()
    default_sheet = result_file_handler.active
    result_file_handler.remove(default_sheet)

    # 收集所有评分结果
    all_results = []
    for name, member in MERITS_MAP.items():
        all_results.append(member.return_excel())

    # 按总分从高到低排序
    all_results.sort(key=lambda x: x[9], reverse=True)

    # 创建统一的评分表
    result_sheet = result_file_handler.create_sheet("评分汇总")
    headers = ['职能组', '考评对象', '员工层级', '备注', '处长评价', '本组组长评价', '其他组组长评价', '本组人员评价', '他组人员评价', '总分']
    result_sheet.append(headers)

    for result in all_results:
        # result 格式：[group, name, level, None, 处长, 本组组长(或0), 其他组组长, 本组人员, 他组人员, 总分]
        # 直接添加完整的结果行
        result_sheet.append(result)

    # 调整列宽
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        result_sheet.column_dimensions[col].width = 15

    result_file_handler.save(RESULT_FILE_PATH)
    logging.info(f"输出文件已生成: {RESULT_FILE_PATH}")


if __name__ == "__main__":
    init()
    init_config()
    check_file()
    computed()
    out()
    input("任务完成。按 Enter 键退出...")

